"""
Sistema de Automação - Tudo do Momento Presentes
Gera relatórios automáticos e planilhas Excel de controle de vendas
"""

import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from src.models.user import db
from src.models.product import Product
from src.models.sale import Sale
from src.models.visit import Visit
from flask import Flask

# Configurar Flask app para acesso ao banco de dados
app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{os.path.join(os.path.dirname(__file__), 'database', 'app.db')}"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db.init_app(app)


def generate_sales_excel_report(period='monthly'):
    """
    Gerar planilha Excel de controle de vendas
    
    Args:
        period: 'weekly' ou 'monthly'
    """
    with app.app_context():
        try:
            # Calcular período
            now = datetime.utcnow()
            if period == 'weekly':
                start_date = now - timedelta(days=7)
                filename = f"vendas_semanal_{now.strftime('%Y%m%d')}.xlsx"
            else:  # monthly
                start_date = now - timedelta(days=30)
                filename = f"vendas_mensal_{now.strftime('%Y%m%d')}.xlsx"
            
            # Buscar vendas do período
            sales = Sale.query.filter(Sale.sale_date >= start_date).order_by(Sale.sale_date.desc()).all()
            
            # Criar workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Controle de Vendas"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Cabeçalho
            headers = [
                "Data da Venda",
                "Nome do Produto",
                "Valor da Venda (R$)",
                "Valor de Compra (R$)",
                "Lucro (R$)",
                "Porcentagem de Lucro (%)",
                "Nome do Comprador",
                "CPF/CNPJ do Comprador"
            ]
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Dados
            for row_idx, sale in enumerate(sales, start=2):
                ws.cell(row=row_idx, column=1, value=sale.sale_date.strftime('%d/%m/%Y %H:%M'))
                ws.cell(row=row_idx, column=2, value=sale.product_name)
                ws.cell(row=row_idx, column=3, value=sale.sale_price)
                ws.cell(row=row_idx, column=4, value=sale.purchase_price)
                ws.cell(row=row_idx, column=5, value=sale.profit)
                ws.cell(row=row_idx, column=6, value=sale.profit_percentage)
                ws.cell(row=row_idx, column=7, value=sale.buyer_name)
                ws.cell(row=row_idx, column=8, value=sale.buyer_document)
                
                # Aplicar bordas
                for col in range(1, 9):
                    ws.cell(row=row_idx, column=col).border = border
            
            # Ajustar largura das colunas
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 22
            ws.column_dimensions['G'].width = 25
            ws.column_dimensions['H'].width = 20
            
            # Adicionar resumo
            summary_row = len(sales) + 3
            ws.cell(row=summary_row, column=1, value="RESUMO DO PERÍODO").font = Font(bold=True, size=14)
            
            total_sales = len(sales)
            total_revenue = sum(sale.sale_price for sale in sales)
            total_profit = sum(sale.profit for sale in sales)
            
            ws.cell(row=summary_row + 1, column=1, value="Total de Vendas:")
            ws.cell(row=summary_row + 1, column=2, value=total_sales)
            
            ws.cell(row=summary_row + 2, column=1, value="Receita Total:")
            ws.cell(row=summary_row + 2, column=2, value=f"R$ {total_revenue:.2f}")
            
            ws.cell(row=summary_row + 3, column=1, value="Lucro Total:")
            ws.cell(row=summary_row + 3, column=2, value=f"R$ {total_profit:.2f}")
            
            # Salvar arquivo
            reports_dir = os.path.join(os.path.dirname(__file__), 'reports')
            os.makedirs(reports_dir, exist_ok=True)
            filepath = os.path.join(reports_dir, filename)
            wb.save(filepath)
            
            print(f"Planilha gerada com sucesso: {filepath}")
            return filepath
            
        except Exception as e:
            print(f"Erro ao gerar planilha Excel: {str(e)}")
            return None


def generate_text_report(period='weekly'):
    """
    Gerar relatório em texto
    
    Args:
        period: 'weekly' ou 'monthly'
    """
    with app.app_context():
        try:
            # Calcular período
            now = datetime.utcnow()
            if period == 'weekly':
                start_date = now - timedelta(days=7)
                period_name = "Semanal"
            else:  # monthly
                start_date = now - timedelta(days=30)
                period_name = "Mensal"
            
            # Buscar dados
            sales = Sale.query.filter(Sale.sale_date >= start_date).all()
            total_sales = len(sales)
            total_revenue = sum(sale.sale_price for sale in sales)
            total_profit = sum(sale.profit for sale in sales)
            
            # Produtos mais vendidos
            from sqlalchemy import func
            top_products = db.session.query(
                Sale.product_name,
                func.count(Sale.id).label('sales_count')
            ).filter(Sale.sale_date >= start_date)\
             .group_by(Sale.product_name)\
             .order_by(func.count(Sale.id).desc())\
             .limit(5)\
             .all()
            
            # Produtos com estoque baixo
            low_stock = Product.query.filter(
                Product.is_active == True,
                Product.stock <= 5
            ).all()
            
            # Estatísticas de visitas
            total_visits = Visit.query.count()
            period_visits = Visit.query.filter(Visit.visit_date >= start_date).count()
            
            # Gerar relatório
            report = f"""
{'=' * 80}
RELATÓRIO {period_name.upper()} - TUDO DO MOMENTO PRESENTES
Data de Geração: {now.strftime('%d/%m/%Y %H:%M:%S')}
Período: {start_date.strftime('%d/%m/%Y')} a {now.strftime('%d/%m/%Y')}
{'=' * 80}

RESUMO FINANCEIRO
-----------------
Total de Vendas: {total_sales}
Receita Total: R$ {total_revenue:.2f}
Lucro Total: R$ {total_profit:.2f}
Margem de Lucro Média: {(total_profit / total_revenue * 100) if total_revenue > 0 else 0:.2f}%

PRODUTOS MAIS VENDIDOS
----------------------
"""
            for idx, (product_name, sales_count) in enumerate(top_products, 1):
                report += f"{idx}. {product_name} - {sales_count} vendas\n"
            
            report += f"""
ALERTAS DE ESTOQUE
------------------
Produtos com estoque baixo (≤5 unidades): {len(low_stock)}
"""
            if low_stock:
                for product in low_stock:
                    report += f"  - {product.name}: {product.stock} unidades\n"
            else:
                report += "  Nenhum produto com estoque baixo.\n"
            
            report += f"""
ESTATÍSTICAS DE VISITAS
-----------------------
Total de Visitas (Geral): {total_visits}
Visitas no Período: {period_visits}

{'=' * 80}
Relatório gerado automaticamente pelo sistema.
© 2025 Tudo do Momento Presentes
{'=' * 80}
"""
            
            # Salvar relatório
            reports_dir = os.path.join(os.path.dirname(__file__), 'reports')
            os.makedirs(reports_dir, exist_ok=True)
            filename = f"relatorio_{period}_{now.strftime('%Y%m%d_%H%M%S')}.txt"
            filepath = os.path.join(reports_dir, filename)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(report)
            
            print(f"Relatório gerado com sucesso: {filepath}")
            return report, filepath
            
        except Exception as e:
            print(f"Erro ao gerar relatório: {str(e)}")
            return None, None


def send_report_email(report_text):
    """
    Enviar relatório por e-mail
    
    Args:
        report_text: Texto do relatório
    """
    try:
        from src.routes.contact import send_email
        
        email_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif;">
            <h2 style="color: #4472C4;">Relatório Automático - Tudo do Momento Presentes</h2>
            <pre style="background-color: #f5f5f5; padding: 15px; border-radius: 5px; overflow-x: auto;">
{report_text}
            </pre>
            <p style="color: #666; font-size: 12px;">
                Este é um relatório automático gerado pelo sistema.
            </p>
        </body>
        </html>
        """
        
        send_email(
            'contato@tudodomomentopresentes.com.br',
            f'Relatório Automático - {datetime.utcnow().strftime("%d/%m/%Y")}',
            email_body
        )
        
        print("Relatório enviado por e-mail com sucesso!")
        return True
        
    except Exception as e:
        print(f"Erro ao enviar relatório por e-mail: {str(e)}")
        return False


def run_weekly_automation():
    """Executar automação semanal completa"""
    print("\n" + "=" * 80)
    print("AUTOMAÇÃO SEMANAL - TUDO DO MOMENTO PRESENTES")
    print(f"Data: {datetime.utcnow().strftime('%d/%m/%Y %H:%M:%S')}")
    print("=" * 80 + "\n")
    
    # 1. Gerar planilha Excel
    print("1. Gerando planilha Excel de vendas...")
    excel_path = generate_sales_excel_report(period='weekly')
    
    # 2. Gerar relatório em texto
    print("\n2. Gerando relatório em texto...")
    report_text, report_path = generate_text_report(period='weekly')
    
    # 3. Enviar relatório por e-mail
    if report_text:
        print("\n3. Enviando relatório por e-mail...")
        send_report_email(report_text)
    
    # 4. Executar pesquisa de mercado
    print("\n4. Executando pesquisa de mercado com IA...")
    try:
        from src.ai_market_research import run_weekly_market_research
        run_weekly_market_research()
    except Exception as e:
        print(f"Erro ao executar pesquisa de mercado: {str(e)}")
    
    print("\n" + "=" * 80)
    print("AUTOMAÇÃO SEMANAL CONCLUÍDA!")
    print("=" * 80 + "\n")


def run_monthly_automation():
    """Executar automação mensal completa"""
    print("\n" + "=" * 80)
    print("AUTOMAÇÃO MENSAL - TUDO DO MOMENTO PRESENTES")
    print(f"Data: {datetime.utcnow().strftime('%d/%m/%Y %H:%M:%S')}")
    print("=" * 80 + "\n")
    
    # 1. Gerar planilha Excel
    print("1. Gerando planilha Excel de vendas...")
    excel_path = generate_sales_excel_report(period='monthly')
    
    # 2. Gerar relatório em texto
    print("\n2. Gerando relatório em texto...")
    report_text, report_path = generate_text_report(period='monthly')
    
    # 3. Enviar relatório por e-mail
    if report_text:
        print("\n3. Enviando relatório por e-mail...")
        send_report_email(report_text)
    
    print("\n" + "=" * 80)
    print("AUTOMAÇÃO MENSAL CONCLUÍDA!")
    print("=" * 80 + "\n")


if __name__ == '__main__':
    import sys
    
    if len(sys.argv) > 1:
        if sys.argv[1] == 'weekly':
            run_weekly_automation()
        elif sys.argv[1] == 'monthly':
            run_monthly_automation()
        else:
            print("Uso: python automation.py [weekly|monthly]")
    else:
        # Executar semanal por padrão
        run_weekly_automation()

